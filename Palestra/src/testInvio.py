import json
from openpyxl import load_workbook
import pandas as pd
import smtplib, os
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import getpass
import win32com.client as win32
import sys
import os



def outlook(email_mittente = 'otto.fontanesi@prometeia.com',
            email_destinatari = [''],
            email_copia_conoscenza = [''],
            email_oggetto = '',
            email_body = '',
            attachment_location = '',
            sending_mode = 'Display'):
    
    ol_app = win32.Dispatch('Outlook.Application')
    ol_ns = ol_app.GetNameSpace('MAPI')
    mail_item = ol_app.CreateItem(0)
    mail_item.Subject = email_oggetto
    mail_item.BodyFormat = 2
    mail_item.Body = email_body

    #mail_item.To = "otto.fontanesi@prometeia.com"
    delim=';'
    mail_item.To = delim.join(email_destinatari)
    mail_item.CC = delim.join(email_copia_conoscenza)
    
    mail_item._oleobj_.Invoke(*(64209, 0, 8, 0, ol_ns.Accounts.Item(email_mittente)))
    
    if sending_mode == "Display":
        mail_item.Display()
        answer = (input("Continuare con la prossima mail se presente (Y/N):"))
        if str(answer) == 'N':
            print(exit)
            sys.exit()
        elif str(answer) == 'Y':
            print('Sto elaborando una nuova mail di richiesta...')
        else:
            #TODO rifare la domanda per adesso break
            print(exit)
            sys.exit()
    elif sending_mode == "Send":
        mail_item.Send()
    else:
        print("Errore nella modalità di invio settata. Scegliere tra 'Display' o 'Send'")
        exit()

def carica_data_frame(nome_file):
    #riferimenti del file excel
    COLONNA_FINALE = 50
    COLONNA_INIZIALE = 2
    RIGA_INTESTAZIONI = 4
    RIGA_INIZIALE = 5
    RIGA_FINALE = 79

    #caricamento file Excel delle richieste
    wb = load_workbook(nome_file, data_only=True) 
    #selezione foglio 01_MAIL
    ws1 = wb["01_MAIL"]

    #creazione e popolamento lista contenente le righe del file con il 
    # codice del centro di costo, il manager, le GU richieste
    #  esattamente nella stessa forma del file Excel
    data = []
    for row in ws1.iter_rows(min_col=COLONNA_INIZIALE, min_row=RIGA_INIZIALE,
                             max_col=COLONNA_FINALE,max_row=RIGA_FINALE,
                             values_only=True):
        data.append(row)
    
    #creazione e popolamento lista contenente 
    # le intestazioni presenti sul file Excel
    header = []
    for row in ws1.iter_rows(min_col=COLONNA_INIZIALE,min_row=RIGA_INTESTAZIONI, 
                             max_col=COLONNA_FINALE, max_row=RIGA_INTESTAZIONI, 
                             values_only=True):
        for element in row:
            header.append(element)
    #creazione del dataframe Pandas filtrando le righe in
    #  cui complessivamente non ci sono richieste per il periodo di interesse
    df_gu = pd.DataFrame(data, columns=header)
    df_gu = df_gu[df_gu['GU_Totale']>0]
    return df_gu

def send_email(email_receiver,
               email_subject,
               email_message,
               attachment_location = '',
               email_sender = 'otto.fontanesi@prometeia.com'):

    msg = MIMEMultipart()
    
    msg['From'] = email_sender
    msg['To'] = email_receiver
    msg['Subject'] = email_subject
    print(email_message)
    msg.attach(MIMEText(email_message, 'plain'))

    if attachment_location != '':
        filename = os.path.basename(attachment_location)
        attachment = open(attachment_location, "rb")
        part = MIMEBase('application', 'octet-stream')
        part.set_payload(attachment.read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition',
                        "attachment; filename= %s" % filename)
        msg.attach(part)

    try:
        server = smtplib.SMTP(host="smtp.office365.com", port=587, timeout=20)
        server.starttls()
        server.ehlo()
        password = getpass.getpass(f"Per favore, inserisci la password per l'indirizzo mail {email_sender}:\n")
        server.login(email_sender, password)
        text = msg.as_string()
        server.sendmail(email_sender, email_receiver, text)
        print(f'email sent to {email_receiver}')
        server.quit()

    except Exception as err:
        print(f"SMPT server connection error for receiver {email_receiver}. Unexpected error:{err}")

    return True


def invio_richieste(df_gu, anagrafica_team, account):
    for ind in df_gu.index:
        copia_conoscenza =[]
        destinatari =[df_gu['mail destinatario principale'][ind]]
        #creazione lista destinatari
        secondo_destinatario = str(df_gu['Secondo destinatario'][ind])
        if ("@prometeia.com" in secondo_destinatario):
            destinatari.append(secondo_destinatario)
        else: 
            print('secondo destinatario assente o scritto male')
        #creazione del messaggio
        periodo = df_gu['Periodo'][ind]
        msg = "Ciao {},\n la presente per richiedere le giornate per la settimana {} per i diversi componenti del team:\n".format(df_gu['Nome Manager'][ind], periodo)
        
        oggetto = df_gu['Oggetto'][ind]
        # creazione elenco puntato per singola persona
        for userId, info in anagrafica_team.items():
            try:
                GU = float(df_gu['GU_'+ userId][ind])
                if GU > 0.0:
                    descrizione_attivita = df_gu["DES_" + userId][ind]
                     #nome e cognome
                    nc  =" ".join([info['nome'],info['cognome']])+" ("+info['ruolo']+")"
                    #numero GU e descrizione attività
                    ric = " ".join([str(GU) , descrizione_attivita]) 
                    #stringa che racchiude il nome e cognome,
                    #  l'attività e le GU richieste
                    richiesta = "\t - {}: {}\n".format(nc, ric) 
                    msg += richiesta
                    copia_conoscenza.append(info['email'])
                #chiusura messaggio
            except Exception as e:
                print('Problemi nella generazione delle richieste per {}. Eccezione: {} \n\nInterruzione della procedura'.format(userId,e))
                sys.exit()
        msg += "\nGrazie mille."
                
        outlook(email_mittente = account,
                email_destinatari = destinatari,
                email_copia_conoscenza = copia_conoscenza,
                email_oggetto = oggetto,
                email_body = msg,
                attachment_location = '',
                sending_mode = 'Display')


def main():    
    sep="\n"+ 100*'*'
    src_path = os.path.dirname(__file__)
    config_path = os.path.realpath(os.path.join(src_path, '..', 'config'))
    anagrafica_team = json.load(open(config_path+'\\anagrafica_team.json'))
    param = json.load(open(config_path+'\\param.json'))
    file_caricato = param['one_drive_path']+param['file_richieste']
    print(sep)
    print('Sto caricando il file '+file_caricato)
    print(sep)
    df_gu = carica_data_frame(file_caricato)
    account = param['email']
    print(sep)
    print('email del mittente: '+account)
    print("inizio l'elaborazione delle richieste.")
    print(sep)
    invio_richieste(df_gu, anagrafica_team, account)

if __name__ == "__main__":
    main()

